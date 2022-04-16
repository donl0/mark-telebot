import asyncio

from openpyxl import load_workbook

from ..utils.base_messages import msg_10, msg_11, msg_25, msg_61, msg_60, msg_36, msg_35, msg_34, msg_38, msg_37, \
    msg_33, msg_32, msg_31, msg_30, msg_29, msg_28, msg_27, msg_26, msg_24, msg_23, msg_22, msg_21, msg_20, msg_19, \
    msg_18, msg_17, msg_15, msg_14, msg_13, msg_12, msg_16, msg_39, msg_40, msg_41, msg_42, msg_70, msg_71, msg_72, \
    msg_73, msg_51, msg_54, msg_52, msg_50, msg_53, msg_81, msg_80, msg_92, msg_93
from ..utils.dbcommands import user_start_update, get_message, user_find_async, type_of_job, user_find_from_jоb_with_id, \
    mark_someone, get_collegaues_from_user, get_employers_from_user, get_teamleaders_from_user, \
    remove_from_manyomany_colleagues, remove_from_manyomany_teamLeader, remove_from_manyomany_employers, \
    increase_to_count, increase_counter_marked_user, mark_yourself, end_or_not_marking_after_yourself, many_mark_me, \
    user_get_many_marked_me, user_get_mark_myself, get_hr_ids, marked_by_someone, user_report_is, \
    get_employers_list_with_rafy_report, user_exist_check_with_dict, create_user_to_AllEmployment, \
    add_to_user_teamLeader, add_to_user_empl, add_to_user_coll, get_hr
from ..utils.keyboards import phone_keyboard, hr_keyboard, teamleader_mark_keyboard, \
    specialist_mark_keyboard, manager_mark_keyboard, top_manager_mark_keyboard, mark_keyboard, mark_comment_keyboard, \
    mark_colleagues_keyboard, mark_temleader_keyboard, mark_employer_keyboard, rady_keyboard, HR_start_keyboard
from ..utils.states import OrderDataUser, FSMContext, State
from aiogram import Bot, Dispatcher
from aiogram import types
from aiogram.types import ReplyKeyboardRemove, \
    ReplyKeyboardMarkup, KeyboardButton, \
    InlineKeyboardMarkup, InlineKeyboardButton, ContentType

from ..utils.text import telegram_markup
from ..utils.tools import plus_all_from_different_models, get_only_names_from_users


async def document_handlers(bot: Bot, dp: Dispatcher):
    @dp.message_handler(content_types=ContentType.DOCUMENT)
    async def get_excel_to_add_users_HR(message: types.Message, state: FSMContext):
        id_person = message['from']['id']
        dict = {'id_tele': id_person, 'HR':True}
       # user_job = await type_of_job(dict)
        if await get_hr(dict):
            print('start excel scan')
            await bot.download_file_by_id(message['document']['file_id'], 'excel_users_informtion_last.xlsx')

            wb = load_workbook('excel_users_informtion_last.xlsx')

            for ccel_row in range(1,300):
                ws = wb["Контакты"]
                user_name = ws['A'+str(ccel_row)].value
                if user_name != None:
                    #если строка не пустая ( с юзером0 проверяем есть ли он у нас в боте)
                    dict = {'fio': user_name}
                    if not await user_exist_check_with_dict(dict):
                        #нужно найи должнось

                        fio=user_name
                        phone_num=ws['B' + str(ccel_row)].value
                        ws = wb["База данных"]
                        for ccel_row in range(1,300):
                            if ws['B'+str(ccel_row)].value == user_name:
                                job_type=ws['C'+str(ccel_row)].value

                                if job_type == 'Специалист':
                                    job_type = 'specialist'
                                elif job_type == 'Менеджер':
                                    job_type = 'manager'
                                elif job_type == 'Топ-Менеджер':
                                    job_type = 'TopManager'
                                dict = {'fio': fio, 'phone_num': phone_num, 'job_type':job_type}
                                await create_user_to_AllEmployment(dict)
                                break
            #нужен цикл 300 по ячейкам B сам человек, D E рук, [F, G, HI, J,K , L, M] КОЛЛ, [N, O, P, Q, R, S, T, U] ПОДЧ
            for ccel_row in range(3,300):
                ws = wb["База данных"]
                user_name = ws['B' + str(ccel_row)].value
                if user_name != None:
                    teamlead1 = ws['D' + str(ccel_row)].value
                    if teamlead1 != None:
                        try:
                            await add_to_user_empl(teamlead1,user_name)
                        except:
                            pass
                    teamlead2 = ws['E' + str(ccel_row)].value
                    if teamlead2 != None:
                        try:
                            await add_to_user_empl(teamlead2,user_name)
                        except:
                            pass
                    collegues_columns = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
                    for column in collegues_columns:
                        colleague = ws[column + str(ccel_row)].value
                        if colleague != None:
                            try:
                                await add_to_user_coll(colleague, user_name)
                            except:
                                pass
                    empl_columns = ['N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
                    for column in empl_columns:
                        empl = ws[column + str(ccel_row)].value
                        if empl != None:
                            try:
                                await add_to_user_teamLeader(empl, user_name)
                            except:
                                pass