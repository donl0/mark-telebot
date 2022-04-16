from openpyxl import Workbook
from openpyxl import load_workbook

from BotShell.utils.dbcommands import get_message, get_message_no_async, get_user_fio
from BotShell.utils.text import telegram_markup


async def excel_make(dict, phone_num, who_is_marking, job_type):
    '''if not is_marked:
        wb = load_workbook('excelFiles/original.xlsx')
        ws = wb.active
    else:
        wb = load_workbook(f'excelFiles/{phone_num}.xlsx')
        ws = wb.active'''
    #Manager'

  #Specialist'
  #'TopManager'
    print('--------------------------------')
    print(job_type)
    if job_type=='TopManager' or job_type=='Manager':
        print('ТУТ1')
        if who_is_marking=='colleg':
            column='O'
        elif who_is_marking=='me':
            column = 'F'
        elif who_is_marking=='teamLead':
            column = 'L'
        elif who_is_marking=='employer':
            column = 'I'


        dict_for_search={'phone_num':phone_num}
        fio=await get_user_fio(dict_for_search)
        wb = load_workbook(f'excelFiles/{fio}.xlsx')
        ws = wb["Менеджер"]

        i = 7
        for mark in dict['answers']:
            if mark==None:
                i += 1
                continue
            ccel = column + str(i)
            print(ccel)
            ccel_for_value = ws[ccel]
            if ccel_for_value.value==None:
                ws[ccel] = int(mark)
            else:
                try:
                    ccel = column + str(i)
                    value_ccel = int(ws[ccel].value)
                    mark = (value_ccel + int(mark)) / 2
                    ws[ccel] = float(mark)
                except:
                    mark = 2
                    ccel = column + str(i)
                    value_ccel = int(ws[ccel].value)
                    mark = (value_ccel + int(mark)) / 2
                    ws[ccel] = float(mark)
            i += 1

        mass_ccels = ['T7', 'T13', 'T18', 'T24', 'T29']
        num=0
        for ccel1 in mass_ccels:
            ccel_for_value = ws[ccel1]
            if ccel_for_value.value==None:
                if dict['comments'][num]!='':
                    ws[ccel1]=dict['comments'][num]
            else:

                value_ccel = ws[ccel1]
                ws[ccel1] = value_ccel.value + dict['comments'][num]
            num+=1

        #wb.save(f'excelFiles/{fio}.xlsx')


        if job_type == 'TopManager':
            print('ТУТ2')
            ws = wb["data"]
            mass_let = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                        'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']

            if who_is_marking == 'colleg':
                column = 'C'
                for row in range(8, 16):

                    ccel_for_value = ws[column + str(row)]
                    if ccel_for_value.value == None:
                        break
                row2 = row+20
            elif who_is_marking == 'me':
                row = 4
                row2 = 26
            elif who_is_marking == 'teamLead': #оцениваю лидера своего
                column = 'C'
                for row in range(17, 25):

                    ccel_for_value = ws[column + str(row)]
                    if ccel_for_value.value == None:
                        break
                row2=row+20
            elif who_is_marking == 'employer': #оцениваю подчин
                row = 6
                row2 = 27

            mark_counter = 0
            comm_counter = 0
            for column in mass_let:
                if column == 'I' or column == 'O' or column == 'O' or column == 'V' or column == 'AB' or column == 'AI':
                    if dict['comments'][comm_counter] == '':
                        pass
                    else:
                        ws[column + str(row)] = dict['comments'][comm_counter]
                    comm_counter += 1
                else:
                    if dict['answers'][mark_counter] == None:
                        pass
                    else:
                        ws[column + str(row)] = int(dict['answers'][mark_counter])
                    mark_counter += 1

            mark_counter = 0
            comm_counter = 0
            for column in mass_let:
                if column == 'I' or column == 'O' or column == 'O' or column == 'V' or column == 'AB' or column == 'AI':
                    if dict['comments'][comm_counter] == '':
                        pass
                    else:
                        ws[column + str(row2)] = dict['comments'][comm_counter]
                    comm_counter += 1
                else:
                    if dict['answers'][mark_counter] == None:
                        pass
                    else:
                        ws[column + str(row2)] = int(dict['answers'][mark_counter])
                    mark_counter += 1
        elif job_type == 'Manager':
            print('ТУТ3')
            ws = wb["data"]
            mass_let = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                        'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']

            if who_is_marking == 'colleg':
                column = 'C'
                for row in range(6, 14):

                    ccel_for_value = ws[column + str(row)]
                    if ccel_for_value.value == None:
                        break
                row2 = row+20
            elif who_is_marking == 'me':
                row = 4
                row2 = 24
            elif who_is_marking == 'teamLead': #оцениваю лидера своего
                column = 'C'
                for row in range(15, 23):

                    ccel_for_value = ws[column + str(row)]
                    if ccel_for_value.value == None:
                        break
                row2=row+20
            elif who_is_marking == 'employer': #оцениваю подчин
                row = 5
                row2 = 25

            mark_counter = 0
            comm_counter = 0
            for column in mass_let:
                if column == 'I' or column == 'O' or column == 'O' or column == 'V' or column == 'AB' or column == 'AI':
                    if dict['comments'][comm_counter] == '':
                        pass
                    else:
                        ws[column + str(row)] = dict['comments'][comm_counter]
                    comm_counter += 1
                else:
                    if dict['answers'][mark_counter] == None:
                        pass
                    else:
                        ws[column + str(row)] = int(dict['answers'][mark_counter])
                    mark_counter += 1

            mark_counter = 0
            comm_counter = 0
            for column in mass_let:
                if column == 'I' or column == 'O' or column == 'O' or column == 'V' or column == 'AB' or column == 'AI':
                    if dict['comments'][comm_counter] == '':
                        pass
                    else:
                        ws[column + str(row2)] = dict['comments'][comm_counter]
                    comm_counter += 1
                else:
                    if dict['answers'][mark_counter] == None:
                        pass
                    else:
                        ws[column + str(row2)] = int(dict['answers'][mark_counter])
                    mark_counter += 1
        wb.save(f'excelFiles/{fio}.xlsx')
    elif job_type=='Specialist':

        if who_is_marking=='colleg':
            column='L'
        elif who_is_marking=='me':
            column = 'F'

        elif who_is_marking=='employer':
            column = 'I'

        dict_for_search = {'phone_num': phone_num}
        fio = await get_user_fio(dict_for_search)
        wb = load_workbook(f'excelFiles/{fio}.xlsx')
        ws = wb["Специалист"]

        i = 10
        for mark in dict['answers']:
            if mark == None:
                i += 1
                continue
            ccel = column + str(i)
            ccel_for_value = ws[ccel]
            if ccel_for_value.value == None:
                ws[ccel] = int(mark)
            else:
                try:
                    ccel = column + str(i)
                    value_ccel = int(ws[ccel].value)
                    mark = (value_ccel + int(mark)) / 2
                    ws[ccel] = float(mark)
                except:
                    mark = 2
                    ccel = column + str(i)
                    value_ccel = int(ws[ccel].value)
                    mark = (value_ccel + int(mark)) / 2
                    ws[ccel] = float(mark)
            i += 1

        mass_ccels = ['Q10', 'Q16', 'Q21', 'Q27', 'Q32']
        num = 0
        for ccel1 in mass_ccels:
            ccel_for_value = ws[ccel1]
            if ccel_for_value.value == None:
                if dict['comments'][num] != '':
                    ws[ccel1] = dict['comments'][num]
            else:

                value_ccel = ws[ccel1]
                ws[ccel1] = value_ccel.value + dict['comments'][num]
            num += 1

        if job_type == 'Specialist':
            print('ТУТ5')
            ws = wb["data"]
            mass_let = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                        'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']

            if who_is_marking == 'colleg':
                column = 'C'
                for row in range(8, 16):

                    ccel_for_value = ws[column + str(row)]
                    if ccel_for_value.value == None:
                        break
                row2 = row+10
            elif who_is_marking == 'me':
                row = 4
                row2 = 16

            elif who_is_marking == 'employer': #оцениваю подчин
                row = 6
                row2 = 17

            mark_counter = 0
            comm_counter = 0
            for column in mass_let:
                if column == 'I' or column == 'O' or column == 'O' or column == 'V' or column == 'AB' or column == 'AI':
                    if dict['comments'][comm_counter] == '':
                        pass
                    else:
                        ws[column + str(row)] = dict['comments'][comm_counter]
                    comm_counter += 1
                else:
                    if dict['answers'][mark_counter] == None:
                        pass
                    else:
                        ws[column + str(row)] = int(dict['answers'][mark_counter])
                    mark_counter += 1

            mark_counter = 0
            comm_counter = 0
            for column in mass_let:
                if column == 'I' or column == 'O' or column == 'O' or column == 'V' or column == 'AB' or column == 'AI':
                    if dict['comments'][comm_counter] == '':
                        pass
                    else:
                        ws[column + str(row2)] = dict['comments'][comm_counter]
                    comm_counter += 1
                else:
                    if dict['answers'][mark_counter] == None:
                        pass
                    else:
                        ws[column + str(row2)] = int(dict['answers'][mark_counter])
                    mark_counter += 1

        wb.save(f'excelFiles/{fio}.xlsx')