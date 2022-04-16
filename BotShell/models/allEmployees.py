from django.db import models
from django.db.models.signals import m2m_changed, pre_save, post_save
from openpyxl import load_workbook



class MY_CHOICES(models.Model):
    choice = models.CharField(max_length=154, unique=True)

    def __str__(self):
        return self.choice

    class Meta:
        verbose_name = 'Существующие роли'
        verbose_name_plural = 'Существующие роли'


class AllEmployment(models.Model):
    fio = models.CharField(max_length=100, default="NULL", verbose_name="Ф.И.О.", unique=True)
    phone_num = models.CharField(max_length=25, default="+11111111111", verbose_name="Номер телефона", unique=True)
    #colleagues = models.ManyToManyField("self", blank=True, verbose_name="Коллеги")
    role = [
        ('teamLead', 'Руководителя'),
        ('employer', 'Подчиненный'),
        ('colleague', 'Коллега'),
    ]

    #role = models.ManyToManyField(MY_CHOICES, verbose_name="Роль сотрудника", null=True, blank=True)

    #jobs = [
    #    ('specialist', 'Специалист'),
    #    ('manager', 'Менеджер'),
    #    ('TopManager', 'Топ-Менеджер'),
    #    ('HR', 'HR')
    #]

    jobs = [
        ('specialist', 'Специалист'),
        ('manager', 'Менеджер'),
        ('TopManager', 'Топ-Менеджер')
    ]
    job_type = models.CharField(choices=jobs, max_length=25, verbose_name='Должность')
    HR = models.BooleanField(default=False, verbose_name='Является ли эшаром')
    mail = models.CharField(max_length=100, default="NULL", verbose_name="@mail")
    id_tele = models.IntegerField(default=None, verbose_name="Телеграм id", null=True, blank=True)
    name_tele = models.CharField(max_length=50, default="NULL", verbose_name="телеграм имя")

    def __str__(self):
        return self.fio+' '+self.phone_num

    class Meta:
        verbose_name = 'Все Сотрудники'
        verbose_name_plural = 'Все Сотрудники'



def creta_table_on_crete_user(sender, instance, created, **kwargs):
    ###сдлетаь трай жэкс если нету роли
    if created:
        print(instance.job_type)
        print(instance.job_type)
        if instance.job_type == 'manager':
            wb = load_workbook('excelFiles/original_manager.xlsx')
            ws = wb["data"]
            ws['A4']=instance.fio

            wb.save(f'excelFiles/{instance.fio}.xlsx')
        elif instance.job_type == 'specialist':

            wb = load_workbook('excelFiles/original_spec.xlsx')
            ws = wb["data"]
            ws['A4']=instance.fio
            wb.save(f'excelFiles/{instance.fio}.xlsx')
        elif instance.job_type == 'TopManager':
            wb = load_workbook('excelFiles/original_top_manager.xlsx')
            ws = wb["data"]
            ws['A4'] = instance.fio
            wb.save(f'excelFiles/{instance.fio}.xlsx')
 #   user = AllEmployment.objects.get(phone_num=instance.phone_num)

   # role = user.role.all()
   # print(role)
   # print('SDELALA')
   # if 'Руководитель' in role:
       # print('1')
        #TeamLeader.objects.get_or_create(user)
       # add_to_teamleader(user)
   # elif 'Подчиненный' in role:
     #   print('2')
     #   Employment.objects.get_or_create(user)
       # add_to_employment(user)
  #  elif 'Коллега' in role:
       # print('3')
       # Сolleague.objects.get_or_create(user)
      #  add_to_colleague(user)
   # print(role[0])
    #list(user.role.all())
   # print(instance.all())
   # print(user.role.all())
    #print(user)



post_save.connect(creta_table_on_crete_user, sender = AllEmployment)
